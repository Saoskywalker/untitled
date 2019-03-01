import sys, random, threading, time
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import QUrl
from PyQt5.QtWidgets import (QWidget, QPushButton, QMainWindow,
    QHBoxLayout, QVBoxLayout, QApplication)
from PyQt5.QtWebEngineWidgets import QWebEngineView
from qt_gui import Ui_MainWindow

#Windows下独立线程, linux不用
import ctypes
ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID("myappid")

data = []
class MainWindows(QMainWindow, QWidget, Ui_MainWindow):
    def __init__(self):
        super(MainWindows, self).__init__()
        '''
        palette1 = QtGui.QPalette()
        palette1.setColor(self.backgroundRole(), QtGui.QColor(192,253,123))   # 设置背景颜色
        # 设置背景图片
        palette1.setBrush(self.backgroundRole(),
                          QtGui.QBrush(QtGui.QPixmap('timg.jpg')))
        self.setPalette(palette1)
        self.setAutoFillBackground(True) # 不设置也可以
        '''
        self.setupUi(self)
        self.setWindowTitle('demo') #窗口标题
        self.setWindowIcon(QtGui.QIcon('timg.jpg')) #标题栏图标
        self.setWindowFlags(QtCore.Qt.FramelessWindowHint)  # 隐藏边框
        self.setWindowOpacity(1)  # 设置窗口透明度
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)  # 设置窗口背景透明
        #self.centralwidget.setFixedSize(800, 480)   #固定窗口大小
        #self.centralwidget.resize(800, 480) #设置窗口大小
        self.label.hide()
        self.label.hideFlag = True
        self.widget.hide()
        self.widget.hideFlag = True
        self.widget_2.hideFlag = True
        self.widget_2.hide()

        self.TextValue = 0  #初始化变量
        self.pushButton.clicked.connect(self.upValue)   #链接槽函数
        self.pushButton_2.clicked.connect(self.dowmValue)
        self.Connect.clicked.connect(self.textUpdata)

        self.pushButton.setStyleSheet(
            '''
            QPushButton
            {
            background:#F76677;
            border-radius:5px;
            }
            QPushButton:hover{background:pink;
            }
            ''')
        self.pushButton_2.setStyleSheet('''
            QPushButton{
               /* background:#F7D674;
                border-radius:5px;*/
                border:none;
                color:gray;
                font-size:12px;
                height:40px;
                padding-left:5px;
                padding-right:10px;
                text-align:left;
            }
            QPushButton:hover{
                color:black;
                border:1px solid #F3F3F5;
                border-radius:10px;
                background:LightGray;
            }
             ''')
        self.toolButton.setStyleSheet(
            '''
                QToolButton{border:none;}
                QToolButton:hover{border-bottom:2px solid #F76677;}
            ''')
        self.toolButton.setToolTip("""Hide ----------
                           ----------             Window""")  # 设置提示文本
        self.label_4.setStyleSheet('''
            QLabel{
                border:none;
                font-size:12px;
                font-weight:700;
                font-family: "Helvetica Neue", Helvetica, Arial, sans-serif;
            }
            ''')
        self.label_5.setStyleSheet('''
            QLabel{
                color:yellow;
                border:none;
                font-size:32px;
                font-weight:700;
                font-family: "Helvetica Neue", Helvetica, Arial, sans-serif;
            }
            ''')
        self.label_5.setAlignment(QtCore.Qt.AlignCenter)  # 设置布局内部件居中显示
        self.progressBar.setStyleSheet('''
            QProgressBar::chunk {
                background-color: #F76677;
            }
        ''')

        self.progressBar.setValue(49)
        self.progressBar.setFixedHeight(10)  # 设置进度条高度
       # self.progressBar.setTextVisible(False)  # 不显示进度条文字
        self.progressBar.setAlignment(QtCore.Qt.AlignCenter)  # 设置布局内部件居中显示
        self.progressBar.setToolTip("Processing")   #设置提示文本

        #滑条链接进度条
        self.horizontalSlider.valueChanged.connect(self.progressBar.setValue)

        #创建WEB报表
        # self.webview = QWebEngineView()
        # self.webview.load(QUrl("file:///D:/Aysi/Python/qt5/render.html"))

    # 鼠标点击事件
    def mousePressEvent(self, event):
        #super(MainWindows, self).mousePressEvent(event)
        #if event.button()==Qt.LeftButton:
        pos = event.pos()
        if self.label.hideFlag==True:
            self.label.hideFlag = False
            self.label.show()

        if self.label.hideFlag==False:
            if (pos.x() >= 160 and pos.x() <= 260) and (pos.y() >= 100 and pos.y() <= 150):
                write_excel()
                self.close()
            elif (pos.x() >= 20 and pos.x() <= 120) and (pos.y() >= 100 and pos.y() <= 150):
                self.showMinimized()
            elif (pos.x() >= 20 and pos.x() <= 120) and (pos.y() >= 200 and pos.y() <= 250):
                if self.widget_2.hideFlag==True:
                    self.widget_2.hideFlag = False
                    self.widget_2.raise_()  #控件顶层显示
                    self.widget_2.show()
                else:
                    self.widget_2.hideFlag = True
                    self.widget_2.hide()
            elif (pos.x() >= 160 and pos.x() <= 260) and (pos.y() >= 200 and pos.y() <= 250):
                if self.widget.hideFlag==True:
                    self.widget.hideFlag = False
                    self.widget.raise_()    #控件顶层显示
                    self.widget.show()
                else:
                    self.widget.hideFlag = True
                    self.widget.hide()

    #键盘处理
    """
    def keyPressEvent(self, e):
    
    if e.key() == Qt.Key_Escape:
        self.close()
    """

    def upValue(self):
        self.TextValue = self.TextValue+1
        self.label_5.setText(str(self.TextValue))

    def dowmValue(self):
        self.TextValue = self.TextValue-1
        self.label_5.setText(str(self.TextValue))

    def textUpdata(self):
        # y = ''
        # for x in data:
        #     y = y+x+'\n'
        # self.textBrowser.setText(y)
        #for g in range(10):
        if(len(data)==0):
            self.textBrowser.setText("NOT THING") 
            return
        i = random.choice(data)
        j = data.index(i)
        del data[j] 
        self.textBrowser.setText(i)       

#excel operate
from openpyxl import Workbook
from openpyxl import load_workbook

#READ EXCEL
def read_excel():
    # 创建文件对象
    workbook = load_workbook(r'C:\Users\Aysi\PycharmProjects\untitled\ddd.xlsx')
    ws =  workbook.active #get frist sheet
    ws1 = workbook.get_sheet_by_name("ui")
    cols = []
    for col in ws1.iter_cols():
        cols.append(col)
 
    # print cols   #所有列
    print(cols[0])   #获取第一列
    # print cols[0][0]   #获取第一列的第一行的单元格对象
    # print cols[0][0].value   #获取第一列的第一行的值
 
    return cols[0]

def write_excel():
    pass

if __name__ == "__main__":
    app = QApplication(sys.argv)
    widget = MainWindows()
    widget.show()
    data = list(read_excel())
    sys.exit(app.exec())
